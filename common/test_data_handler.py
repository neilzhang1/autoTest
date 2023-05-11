# @Time : 2022/6/9 0:49 
# encoding: utf-8
# @Author : Neil
# @File : test_data_handler.py 
# @Software: PyCharm
"""
测试数据处理
"""
import json
import re

from openpyxl import load_workbook
from faker import Faker
from common import db


def get_test_data_from_excel(file, sheet_name):
    """
    获取excel中的用例数据
    """
    # 1.打开excel
    wb = load_workbook(filename=file)
    # 2.获取sheet
    sheet = wb[sheet_name]
    row = sheet.max_row
    column = sheet.max_column
    # 3.读取数据
    data = []
    # 获取第一行并拿到所有的key
    keys = []
    for i in range(1, column + 1):
        keys.append(sheet.cell(1, i).value)
    # 循环每一行获取数据并组成字典
    for i in range(2, row + 1):
        row_data = {}
        for j in range(1, column + 1):
            # 获取对应列的键
            # key = keys[j-1]
            # value = sheet.cell(i,j).value
            # row_data[key] = value
            row_data[keys[j - 1]] = sheet.cell(i, j).value
        # 把部分数据转换为python对象
        # try:
        #     row_data['request'] = json.loads(row_data['request'])
        #     row_data['expect'] = json.loads(row_data['expect'])
        # except json.decoder.JSONDecodeError:
        #     raise ValueError('用例数据json格式错误')
        # 把每一行的字典添加到列表中
        data.append(row_data)
    return data


def generate_phone():
    """
    随机生成手机号
    """
    fk = Faker(locale='zh_CN')
    return fk.phone_number()


def generate_no_use_phone(sql='select telephone from userInfo where telephone = {} '):
    """
    随机生成没有使用过的手机号码
    """

    while True:
        phone = generate_phone()
        sql = sql.format(phone)
        if not db.exist(sql):
            return phone


def replace_args_by_re(json_data, obj):
    """
    使用正则表达式动态替换参数
    """
    # 1.找出所有槽位中的变量名
    args = re.findall(r'#(.*?)#', json_data)
    # 2.根据变量名找到对应的值
    for arg in args:
        value = getattr(obj, arg, None)
        if value:
            json_data = json_data.replace('#{}#'.format(arg), str(value))

    return json_data


if __name__ == '__main__':
    res = get_test_data_from_excel('../testdata/testcases.xlsx', 'register')
    print(type(res[1]['request']))
